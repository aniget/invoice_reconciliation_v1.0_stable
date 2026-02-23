"""
EVD Data Extractor - Standalone Project

This script extracts invoice data from Zaiavka za plashtane (EVD) Excel files
and structures it for easy comparison with PDF invoices.

Author: EVD Extraction Team
Date: 2026-01-28
"""

import re
from datetime import datetime
from pathlib import Path
import openpyxl
import json
import sys
import logging

# Configure logging
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[logging.FileHandler(
                        "debug.log", encoding="utf-8"), logging.StreamHandler()],
                    force=True
                    )
logger = logging.getLogger(__name__)


class EVDExtractor:
    """
    Extracts structured data from EVD (Expense Verification Document) Excel files.

    The EVD format typically has:
    - Row 12-13: Invoice information section headers
    - Row 14+: Invoice data rows

    Key columns identified:
    - Column B (2): Row number
    - Column D (4): Vendor/Supplier name
    - Column O (15): Invoice number
    - Column AC (29): Invoice date
    - Column AJ (36): Currency
    - Column AN (40): Currency amount
    - Column AU (47): Net amount EUR
    - Column BB (54): VAT amount EUR
    - Column BI (61): Total amount EUR
    """

    # Column mapping (1-based index)
    COLUMNS = {
        'row_num': 2,        # B - Row number
        'vendor': 4,         # D - Vendor/Supplier
        'invoice_num': 15,   # O - Invoice number
        'invoice_date': 29,  # AC - Invoice date
        'currency': 36,      # AJ - Currency
        'currency_amount': 40,  # AN - Amount in original currency
        'net_amount_eur': 47,   # AU - Net amount in EUR
        'vat_amount_eur': 54,   # BB - VAT amount in EUR
        'total_amount_eur': 61,  # BI - Total amount in EUR
    }

    def __init__(self, file_path):
        """
        Initialize extractor with EVD file path.

        Args:
            file_path (str or Path): Path to EVD Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None
        self.data_start_row = 14  # Default starting row for invoice data

    def load_file(self):
        """Load the Excel workbook."""
        logger.info(f"Loading file: {self.file_path}")
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        logger.info(f"Opening file: {self.file_path}")

        self.worksheet = self.workbook.active
        logger.info(f"  Sheet: {self.worksheet.title}")
        logger.info(f"  Dimensions: {self.worksheet.dimensions}")

    def find_data_start_row(self):
        """
        Automatically find where invoice data starts.
        Looks for "Информация за фактура" or "Invoice information".
        """
        for row_idx in range(1, 20):
            cell_val = self.worksheet.cell(row=row_idx, column=2).value
            if cell_val and isinstance(cell_val, str):
                if 'Invoice information' in cell_val or 'Информация за фактура' in cell_val:
                    # Data starts 2 rows after this header
                    self.data_start_row = row_idx + 2
                    logger.info(
                        f"Found data start row: {self.data_start_row}")
                    return

        logger.info(f"Using default data start row: {self.data_start_row}")

    def normalize_vendor_name(self, vendor):
        """
        Normalize vendor name for matching.

        Args:
            vendor (str): Raw vendor name

        Returns:
            str: Normalized vendor name
        """
        if not vendor:
            return ""

        vendor = str(vendor).strip().upper()

        # Remove common company suffixes
        suffixes = ['ЕАД', 'EAD', 'ЕООД', 'EOOD', 'АД',
                    'AD', 'ООД', 'OOD', 'LTD', 'LIMITED', 'LLC']
        for suffix in suffixes:
            vendor = re.sub(rf'\s+{suffix}\s*$', '',
                            vendor, flags=re.IGNORECASE)

        # Remove extra whitespace
        vendor = re.sub(r'\s+', ' ', vendor).strip()

        return vendor

    def extract_invoice_data(self):
        """
        Extract all invoice records from the EVD file.

        Returns:
            list: List of invoice records as dictionaries
        """
        invoices = []

        # Find the last row with data
        max_row = self.worksheet.max_row

        logger.info(
            f"\nExtracting invoice data from rows {self.data_start_row} to {max_row}...")

        for row_idx in range(self.data_start_row, max_row + 1):
            # Check if this row has a vendor (required field)
            vendor = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['vendor']).value

            # Skip empty rows or header rows
            if not vendor or not isinstance(vendor, str) or len(vendor.strip()) < 5:
                continue

            # Skip if it looks like a header or metadata row
            skip_keywords = ['Vendor', 'Доставчик', 'Name', 'Title', 'Function',
                             'Cost center', 'WBS', 'E-mail', 'Phone', 'Date', 'Signature',
                             'Име', 'Длъжност', 'Отдел', 'Разходен', 'СПП', 'Електронна',
                             'Телефон', 'Дата', 'Подпис']
            if any(keyword in vendor for keyword in skip_keywords):
                continue

            # Extract all fields
            row_num = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['row_num']).value
            invoice_num = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['invoice_num']).value
            invoice_date = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['invoice_date']).value
            currency = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['currency']).value
            currency_amount = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['currency_amount']).value
            net_amount_eur = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['net_amount_eur']).value
            vat_amount_eur = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['vat_amount_eur']).value
            total_amount_eur = self.worksheet.cell(
                row=row_idx, column=self.COLUMNS['total_amount_eur']).value

            # Create invoice record
            invoice = {
                'evd_row': row_idx,
                'row_number': row_num,
                'vendor_original': vendor.strip(),
                'vendor_normalized': self.normalize_vendor_name(vendor),
                'invoice_number': str(invoice_num).strip() if invoice_num else '',
                'invoice_date': self._format_date(invoice_date),
                'currency': str(currency).strip() if currency else 'EUR',
                'currency_amount': self._to_float(currency_amount),
                'net_amount_eur': self._to_float(net_amount_eur),
                'vat_amount_eur': self._to_float(vat_amount_eur),
                'total_amount_eur': self._to_float(total_amount_eur),
            }

            # Only add if we have at least invoice number and vendor
            if invoice['invoice_number'] and invoice['vendor_normalized']:
                invoices.append(invoice)
                logger.info(
                    f"  Row {row_idx}: {invoice['vendor_normalized']} - {invoice['invoice_number']} - €{invoice['total_amount_eur']}")

        logger.info(f"\nExtracted {len(invoices)} invoice records")
        return invoices

    def _format_date(self, date_val):
        """Format date to string."""
        if date_val is None:
            return None
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m-%d')
        if isinstance(date_val, str):
            return date_val
        return str(date_val)

    def _to_float(self, value):
        """Convert value to float, handling None and errors."""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def group_by_vendor(self, invoices):
        """
        Group invoices by vendor.

        Args:
            invoices (list): List of invoice dictionaries

        Returns:
            dict: Dictionary with vendor as key and list of invoices as value
        """
        grouped = {}

        for invoice in invoices:
            vendor = invoice['vendor_normalized']
            if vendor not in grouped:
                grouped[vendor] = []
            grouped[vendor].append(invoice)

        return grouped

    def create_comparison_structure(self, invoices):
        """
        Create a structure optimized for comparison with PDF data.

        Args:
            invoices (list): List of invoice dictionaries

        Returns:
            dict: Structured data for comparison
        """
        # Group by vendor
        by_vendor = self.group_by_vendor(invoices)

        # Create comparison structure
        structure = {
            'metadata': {
                'source_file': str(self.file_path.name),
                'extraction_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_invoices': len(invoices),
                'total_vendors': len(by_vendor),
                'total_amount_eur': sum(inv['total_amount_eur'] for inv in invoices)
            },
            'by_vendor': {},
            'by_invoice_number': {},
            'all_invoices': invoices
        }

        # Structure by vendor with summary
        for vendor, vendor_invoices in by_vendor.items():
            structure['by_vendor'][vendor] = {
                'vendor_name': vendor,
                'invoice_count': len(vendor_invoices),
                'total_amount': sum(inv['total_amount_eur'] for inv in vendor_invoices),
                'invoices': vendor_invoices
            }

        # Create invoice number lookup
        for invoice in invoices:
            inv_num = invoice['invoice_number']
            if inv_num:
                if inv_num not in structure['by_invoice_number']:
                    structure['by_invoice_number'][inv_num] = []
                structure['by_invoice_number'][inv_num].append(invoice)

        return structure

    def extract_and_structure(self):
        """
        Main extraction method - loads file and extracts structured data.

        Returns:
            dict: Complete structured data ready for comparison
        """
        self.load_file()
        self.find_data_start_row()
        invoices = self.extract_invoice_data()
        structure = self.create_comparison_structure(invoices)

        return structure

    def save_to_json(self, data, output_path):
        """
        Save extracted data to JSON file.

        Args:
            data (dict): Structured data
            output_path (str or Path): Output file path
        """
        output_path = Path(output_path)
        with open(output_path, 'w', encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"\nSaved to: {output_path}")

    def print_summary(self, data):
        """Print a summary of extracted data."""
        logger.info("\n" + "="*80)
        logger.info("EXTRACTION SUMMARY")
        logger.info("="*80)

        metadata = data['metadata']
        logger.info(f"Source file: {metadata['source_file']}")
        logger.info(f"Extraction date: {metadata['extraction_date']}")
        logger.info(f"Total invoices: {metadata['total_invoices']}")
        logger.info(f"Total vendors: {metadata['total_vendors']}")
        logger.info(
            f"Total amount (EUR): €{metadata['total_amount_eur']:,.2f}")

        logger.info("\nBy Vendor:")
        for vendor, vendor_data in data['by_vendor'].items():
            logger.info(
                f"  {vendor}: {vendor_data['invoice_count']} invoices, €{vendor_data['total_amount']:,.2f}")


def main():
    """Main entry point."""
    logger.info("="*80)
    logger.info("EVD Data Extractor")
    logger.info("="*80)

    if len(sys.argv) < 2:
        logger.info(
            "\nUsage: python evd_extractor.py <evd_file.xlsx> [output.json]")
        logger.info("\nExample:")
        logger.info("  python evd_extractor.py Zaiavka_za_plashtane.xlsx")
        logger.info(
            "  python evd_extractor.py Zaiavka_za_plashtane.xlsx output_data.json")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    # Auto-generate output filename if not provided
    if not output_file:
        input_path = Path(input_file)
        output_file = input_path.stem + '_extracted.json'

    try:
        # Extract data
        logger.info("Start extrating ...")
        extractor = EVDExtractor(input_file)

        data = extractor.extract_and_structure()

        # Print summary
        extractor.print_summary(data)

        # Save to JSON
        extractor.save_to_json(data, output_file)

        logger.info("\n[SUCCESS] Extraction completed successfully!")

    except FileNotFoundError:
        logger.info(f"\n[ERROR] Error: File not found: {input_file}")
        sys.exit(1)
    except Exception as e:
        logger.info(f"\n[ERROR] Error during extraction: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
