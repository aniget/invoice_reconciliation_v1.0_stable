"""
EVD-PDF Reconciliation System

Integrates EVD extraction, PDF extraction, and comparison with Excel report generation.
Produces comprehensive reconciliation reports in Excel format.

Author: Reconciliation Team
Date: 2026-01-29
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys


class ReconciliationReportGenerator:
    """
    Generates professional Excel reconciliation reports.
    Compares EVD data with PDF extractions and highlights discrepancies.
    """

    def __init__(self):
        """Initialize report generator with styling."""
        self.wb = None

        # Define styles
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)

        self.match_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.mismatch_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.missing_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_report(self, evd_data: dict, pdf_data: dict, comparison_results: dict,
                        output_path: Path):
        """
        Generate complete reconciliation report.

        Args:
            evd_data (dict): EVD extraction data
            pdf_data (dict): PDF extraction data
            comparison_results (dict): Comparison results
            output_path (Path): Output Excel file path
        """
        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        print("Generating reconciliation report...")

        # Create sheets
        self._create_summary_sheet(evd_data, pdf_data, comparison_results)
        self._create_matches_sheet(comparison_results)
        self._create_mismatches_sheet(comparison_results)
        self._create_missing_in_pdf_sheet(comparison_results)
        self._create_missing_in_evd_sheet(comparison_results)
        self._create_by_vendor_sheet(comparison_results, evd_data, pdf_data)
        self._create_detailed_comparison_sheet(comparison_results)

        # Save workbook
        self.wb.save(output_path)
        print(f"[SUCCESS] Reconciliation report saved: {output_path}")

    def _create_summary_sheet(self, evd_data: dict, pdf_data: dict, comparison_results: dict):
        """Create summary overview sheet."""
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "EVD-PDF Reconciliation Report"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:D1')

        # Report info
        ws['A3'] = "Report Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if 'source_file' in pdf_data['metadata']:
            ws['A4'] = "Source File"
            ws['B4'] = pdf_data['metadata']['source_file']
        elif 'source_folder' in pdf_data['metadata']:
            ws['A4'] = "Source Folder"
            ws['B4'] = pdf_data['metadata']['source_folder']
        else:
            ws['A4'] = "Source"
            ws['B4'] = "Unknown"

        ws['A5'] = "PDF Files Processed:"
        ws['B5'] = pdf_data['metadata']['total_invoices']

        # Summary statistics
        ws['A7'] = "RECONCILIATION SUMMARY"
        ws['A7'].font = Font(size=12, bold=True)

        summary = comparison_results['summary']

        data = [
            ['Metric', 'Count', 'Percentage'],
            ['Total EVD Invoices', summary['total_evd'], '100%'],
            ['Total PDF Invoices', summary['total_pdf'], '100%'],
            ['Perfect Matches', summary['matches'],
             f"{(summary['matches']/max(summary['total_evd'], 1)*100):.1f}%"],
            ['Mismatches (discrepancies)', summary['mismatches'],
             f"{(summary['mismatches']/max(summary['total_evd'], 1)*100):.1f}%"],
            ['Missing in PDF', summary['missing_in_pdf'],
             f"{(summary['missing_in_pdf']/max(summary['total_evd'], 1)*100):.1f}%"],
            ['Missing in EVD', summary['missing_in_evd'],
             f"{(summary['missing_in_evd']/max(summary['total_pdf'], 1)*100):.1f}%"],
        ]

        # Write data
        row = 8
        for r_data in data:
            for col_idx, value in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)

                if row == 8:  # Header
                    cell.fill = self.header_fill
                    cell.font = self.header_font

                cell.border = self.border
                cell.alignment = Alignment(
                    horizontal='left' if col_idx == 1 else 'center')

            row += 1

        # Color code rows
        ws['B11'].fill = self.match_fill  # Matches
        ws['B12'].fill = self.mismatch_fill  # Mismatches
        ws['B13'].fill = self.missing_fill  # Missing in PDF
        ws['B14'].fill = self.missing_fill  # Missing in EVD

        # Match rate indicator
        match_rate = (summary['matches'] / max(summary['total_evd'], 1)) * 100

        ws['A16'] = "Overall Match Rate:"
        ws['B16'] = f"{match_rate:.1f}%"
        ws['B16'].font = Font(size=14, bold=True,
                              color="006100" if match_rate >= 90 else
                              "9C5700" if match_rate >= 70 else "9C0006")

        # Status
        if match_rate >= 95:
            status = "EXCELLENT - Minimal discrepancies"
            color = "006100"
        elif match_rate >= 85:
            status = "GOOD - Some review needed"
            color = "9C5700"
        else:
            status = "ATTENTION REQUIRED - Significant discrepancies"
            color = "9C0006"

        ws['A17'] = "Status:"
        ws['B17'] = status
        ws['B17'].font = Font(bold=True, color=color)

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_matches_sheet(self, comparison_results: dict):
        """Create sheet for perfect matches."""
        ws = self.wb.create_sheet("Matches")

        # Headers
        headers = ['Vendor', 'Invoice Number', 'EVD Date', 'PDF Date',
                   'EVD Amount (EUR)', 'PDF Amount (EUR)', 'Confidence Score', 'EVD File', 'PDF File']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        # Data
        row = 2
        for match in comparison_results['matches']:
            evd = match['evd']
            pdf = match['pdf']

            ws.cell(row=row, column=1, value=evd['vendor_normalized'])
            ws.cell(row=row, column=2, value=evd['invoice_number'])
            ws.cell(row=row, column=3, value=evd.get('invoice_date', ''))
            ws.cell(row=row, column=4, value=pdf.get('invoice_date', ''))
            ws.cell(row=row, column=5, value=evd['total_amount_eur'])
            ws.cell(row=row, column=6, value=pdf.get('total_amount_eur', 0))
            ws.cell(row=row, column=7, value=match.get('confidence', 100))
            ws.cell(row=row, column=8, value=evd.get('file', ''))
            ws.cell(row=row, column=9, value=pdf.get('filename', ''))

            # Format amounts
            ws.cell(row=row, column=5).number_format = '€#,##0.00'
            ws.cell(row=row, column=6).number_format = '€#,##0.00'

            # Apply border and green fill
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).fill = self.match_fill

            row += 1

        # Auto-size columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.column_dimensions['A'].width = 25  # Vendor
        ws.column_dimensions['H'].width = 35  # EVD File
        ws.column_dimensions['I'].width = 35  # PDF File

    def _create_mismatches_sheet(self, comparison_results: dict):
        """Create sheet for mismatches with discrepancies."""
        ws = self.wb.create_sheet("Mismatches")

        # Headers
        headers = ['Vendor', 'Invoice Number', 'EVD Amount', 'PDF Amount',
                   'Difference', 'Discrepancy Type', 'Details', 'EVD File', 'PDF File']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Data
        row = 2
        for mismatch in comparison_results['mismatches']:
            evd = mismatch['evd']
            pdf = mismatch['pdf']

            ws.cell(row=row, column=1, value=evd['vendor_normalized'])
            ws.cell(row=row, column=2, value=evd['invoice_number'])
            ws.cell(row=row, column=3, value=evd['total_amount_eur'])
            ws.cell(row=row, column=4, value=pdf.get('total_amount_eur', 0))

            # Calculate difference
            diff = abs(evd['total_amount_eur'] -
                       pdf.get('total_amount_eur', 0))
            ws.cell(row=row, column=5, value=diff)

            # Discrepancy details
            discrepancies = mismatch.get('discrepancies', [])
            disc_types = ', '.join([d['type'] for d in discrepancies])
            disc_details = '; '.join(
                [self._format_discrepancy(d) for d in discrepancies])

            ws.cell(row=row, column=6, value=disc_types)
            ws.cell(row=row, column=7, value=disc_details)
            ws.cell(row=row, column=8, value=evd.get('file', ''))
            ws.cell(row=row, column=9, value=pdf.get('filename', ''))

            # Format
            ws.cell(row=row, column=3).number_format = '€#,##0.00'
            ws.cell(row=row, column=4).number_format = '€#,##0.00'
            ws.cell(row=row, column=5).number_format = '€#,##0.00'

            # Apply red fill for mismatches
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).fill = self.mismatch_fill

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 35
        ws.column_dimensions['I'].width = 35

    def _format_discrepancy(self, disc: dict) -> str:
        """Format discrepancy for display."""
        if disc['type'] == 'amount':
            return f"Amount diff: €{disc.get('difference', 0):.2f}"
        elif disc['type'] == 'currency':
            return f"Currency: {disc.get('evd_value')} vs {disc.get('pdf_value')}"
        elif disc['type'] == 'date':
            return f"Date: {disc.get('evd_value')} vs {disc.get('pdf_value')}"
        return str(disc)

    def _create_missing_in_pdf_sheet(self, comparison_results: dict):
        """Create sheet for EVD entries without matching PDF."""
        ws = self.wb.create_sheet("Missing in PDF")

        # Headers
        headers = ['Vendor', 'Invoice Number', 'Invoice Date', 'Amount (EUR)',
                   'Currency', 'Doc Type', 'EVD File', 'Row']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Data
        row = 2
        for evd in comparison_results['missing_in_pdf']:
            ws.cell(row=row, column=1, value=evd['vendor_normalized'])
            ws.cell(row=row, column=2, value=evd['invoice_number'])
            ws.cell(row=row, column=3, value=evd.get('invoice_date', ''))
            ws.cell(row=row, column=4, value=evd['total_amount_eur'])
            ws.cell(row=row, column=5, value=evd.get('currency', 'EUR'))
            ws.cell(row=row, column=6, value=evd.get('doc_type', ''))
            ws.cell(row=row, column=7, value=evd.get('file', ''))
            ws.cell(row=row, column=8, value=evd.get('evd_row', ''))

            # Format
            ws.cell(row=row, column=4).number_format = '€#,##0.00'

            # Yellow fill for missing
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).fill = self.missing_fill

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['G'].width = 40

    def _create_missing_in_evd_sheet(self, comparison_results: dict):
        """Create sheet for PDF files without matching EVD entry."""
        ws = self.wb.create_sheet("Missing in EVD")

        # Headers
        headers = ['Vendor', 'Invoice Number', 'Invoice Date', 'Amount (EUR)',
                   'Currency', 'Confidence', 'PDF File']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Data
        row = 2
        for pdf in comparison_results['missing_in_evd']:
            ws.cell(row=row, column=1, value=pdf.get(
                'vendor_normalized', 'UNKNOWN'))
            ws.cell(row=row, column=2, value=pdf.get('invoice_number', ''))
            ws.cell(row=row, column=3, value=pdf.get('invoice_date', ''))
            ws.cell(row=row, column=4, value=pdf.get('total_amount_eur', 0))
            ws.cell(row=row, column=5, value=pdf.get('currency', 'EUR'))
            ws.cell(row=row, column=6, value=pdf.get('confidence', 0))
            ws.cell(row=row, column=7, value=pdf.get('filename', ''))

            # Format
            ws.cell(row=row, column=4).number_format = '€#,##0.00'

            # Yellow fill
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).fill = self.missing_fill

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['G'].width = 40

    def _create_by_vendor_sheet(self, comparison_results: dict, evd_data: dict, pdf_data: dict):
        """Create vendor summary sheet."""
        ws = self.wb.create_sheet("By Vendor")

        # Headers
        headers = ['Vendor', 'EVD Invoices', 'PDF Invoices', 'Matches',
                   'Mismatches', 'Missing PDF', 'Missing EVD', 'Match Rate']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Get all vendors
        all_vendors = set()
        all_vendors.update(evd_data['by_vendor'].keys())
        all_vendors.update(pdf_data['by_vendor'].keys())

        row = 2
        for vendor in sorted(all_vendors):
            evd_count = len(evd_data['by_vendor'].get(
                vendor, {}).get('invoices', []))
            pdf_count = len(pdf_data['by_vendor'].get(
                vendor, {}).get('invoices', []))

            # Count matches/mismatches for this vendor
            matches = sum(1 for m in comparison_results['matches']
                          if m['evd']['vendor_normalized'] == vendor)
            mismatches = sum(1 for m in comparison_results['mismatches']
                             if m['evd']['vendor_normalized'] == vendor)
            missing_pdf = sum(1 for m in comparison_results['missing_in_pdf']
                              if m['vendor_normalized'] == vendor)
            missing_evd = sum(1 for m in comparison_results['missing_in_evd']
                              if m.get('vendor_normalized') == vendor)

            match_rate = (matches / max(evd_count, 1)) * 100

            ws.cell(row=row, column=1, value=vendor)
            ws.cell(row=row, column=2, value=evd_count)
            ws.cell(row=row, column=3, value=pdf_count)
            ws.cell(row=row, column=4, value=matches)
            ws.cell(row=row, column=5, value=mismatches)
            ws.cell(row=row, column=6, value=missing_pdf)
            ws.cell(row=row, column=7, value=missing_evd)
            ws.cell(row=row, column=8, value=f"{match_rate:.1f}%")

            # Color code based on match rate
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border

                if match_rate >= 90:
                    cell.fill = self.match_fill
                elif match_rate >= 70:
                    cell.fill = self.missing_fill
                else:
                    cell.fill = self.mismatch_fill

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30

    def _create_detailed_comparison_sheet(self, comparison_results: dict):
        """Create detailed side-by-side comparison."""
        ws = self.wb.create_sheet("Detailed Comparison")

        # Headers
        headers = ['Status', 'Vendor', 'Invoice #',
                   'EVD Date', 'PDF Date',
                   'EVD Amount', 'PDF Amount', 'Difference',
                   'EVD Currency', 'PDF Currency',
                   'Notes']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        row = 2

        # Add all matches
        for match in comparison_results['matches']:
            self._add_comparison_row(ws, row, 'MATCH', match['evd'],
                                     match.get('pdf'), self.match_fill)
            row += 1

        # Add mismatches
        for mismatch in comparison_results['mismatches']:
            self._add_comparison_row(ws, row, 'MISMATCH', mismatch['evd'],
                                     mismatch.get('pdf'), self.mismatch_fill,
                                     '; '.join([self._format_discrepancy(d)
                                               for d in mismatch.get('discrepancies', [])]))
            row += 1

        # Add missing in PDF
        for missing in comparison_results['missing_in_pdf']:
            self._add_comparison_row(ws, row, 'NO PDF', missing, None,
                                     self.missing_fill, 'PDF not found')
            row += 1

        # Add missing in EVD
        for missing in comparison_results['missing_in_evd']:
            self._add_comparison_row(ws, row, 'NO EVD', None, missing,
                                     self.missing_fill, 'Not in EVD')
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['K'].width = 50

    def _add_comparison_row(self, ws, row, status, evd, pdf, fill, notes=''):
        """Add a comparison row to detailed sheet."""
        ws.cell(row=row, column=1, value=status)

        if evd:
            ws.cell(row=row, column=2, value=evd.get('vendor_normalized', ''))
            ws.cell(row=row, column=3, value=evd.get('invoice_number', ''))
            ws.cell(row=row, column=4, value=evd.get('invoice_date', ''))
            ws.cell(row=row, column=6, value=evd.get('total_amount_eur', 0))
            ws.cell(row=row, column=9, value=evd.get('currency', 'EUR'))

        if pdf:
            if not evd:  # Only set vendor if not already set
                ws.cell(row=row, column=2, value=pdf.get(
                    'vendor_normalized', ''))
                ws.cell(row=row, column=3, value=pdf.get('invoice_number', ''))
            ws.cell(row=row, column=5, value=pdf.get('invoice_date', ''))
            ws.cell(row=row, column=7, value=pdf.get('total_amount_eur', 0))
            ws.cell(row=row, column=10, value=pdf.get('currency', 'EUR'))

        # Calculate difference
        if evd and pdf:
            diff = abs(evd.get('total_amount_eur', 0) -
                       pdf.get('total_amount_eur', 0))
            ws.cell(row=row, column=8, value=diff)
            ws.cell(row=row, column=8).number_format = '€#,##0.00'

        ws.cell(row=row, column=11, value=notes)

        # Format amounts
        if ws.cell(row=row, column=6).value:
            ws.cell(row=row, column=6).number_format = '€#,##0.00'
        if ws.cell(row=row, column=7).value:
            ws.cell(row=row, column=7).number_format = '€#,##0.00'

        # Apply fill and border
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = self.border


def main():
    """Main execution - load data and generate report."""
    print("="*80)
    print("EVD-PDF Reconciliation System")
    print("="*80)

    if len(sys.argv) < 4:
        print("\nUsage: python reconciliation_report.py <evd_data.json> <pdf_data.json> <output.xlsx>")
        print("\nExample:")
        print("  python reconciliation_report.py evd_extracted.json pdf_extracted.json reconciliation.xlsx")
        sys.exit(1)

    evd_file = Path(sys.argv[1])
    pdf_file = Path(sys.argv[2])
    output_file = Path(sys.argv[3])

    # Load data
    print(f"\nLoading EVD data from: {evd_file}")
    with open(evd_file, 'r', encoding="utf-8") as f:
        evd_data = json.load(f)

    print(f"Loading PDF data from: {pdf_file}")
    with open(pdf_file, 'r', encoding="utf-8") as f:
        pdf_data = json.load(f)

    # Import and use comparator
    from pdf_evd_comparator import EVDPDFComparator

    # Compare
    print("\nComparing datasets...")
    comparator = EVDPDFComparator(amount_tolerance=0.01)
    comparison_results = comparator.compare_datasets(evd_data, pdf_data)

    # Print summary
    comparator.print_comparison_summary(comparison_results)

    # Generate Excel report
    print(f"\nGenerating Excel report...")
    generator = ReconciliationReportGenerator()
    generator.generate_report(
        evd_data, pdf_data, comparison_results, output_file)

    print(f"\n{'='*80}")
    print("RECONCILIATION COMPLETE")
    print(f"{'='*80}")
    print(f"Excel report: {output_file}")
    print(f"\nSummary:")
    print(
        f"  Total EVD invoices: {comparison_results['summary']['total_evd']}")
    print(
        f"  Total PDF invoices: {comparison_results['summary']['total_pdf']}")
    print(f"  Perfect matches: {comparison_results['summary']['matches']}")
    print(f"  Mismatches: {comparison_results['summary']['mismatches']}")
    print(
        f"  Missing in PDF: {comparison_results['summary']['missing_in_pdf']}")
    print(
        f"  Missing in EVD: {comparison_results['summary']['missing_in_evd']}")


if __name__ == "__main__":
    main()
