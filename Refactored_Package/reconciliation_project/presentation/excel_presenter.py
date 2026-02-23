"""
Excel Report Presentation

Pure presentation layer - handles only Excel formatting and styling.
No business logic, no calculations - receives formatted data and presents it.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from datetime import datetime


class ExcelStyles:
    """
    Centralized Excel styling definitions.
    
    All visual styling in one place for easy maintenance.
    """
    
    # Colors (hex codes)
    COLOR_HEADER_BG = "366092"
    COLOR_HEADER_TEXT = "FFFFFF"
    COLOR_MATCH = "C6EFCE"      # Light green
    COLOR_MISMATCH = "FFC7CE"   # Light red
    COLOR_MISSING = "FFEB9C"    # Light yellow
    COLOR_STATUS_GOOD = "006100"
    COLOR_STATUS_WARN = "9C5700"
    COLOR_STATUS_BAD = "9C0006"
    
    # Fonts
    FONT_TITLE = Font(size=16, bold=True, color=COLOR_HEADER_BG)
    FONT_HEADER = Font(color=COLOR_HEADER_TEXT, bold=True, size=11)
    FONT_SECTION = Font(size=12, bold=True)
    FONT_STATUS_GOOD = Font(bold=True, color=COLOR_STATUS_GOOD)
    FONT_STATUS_WARN = Font(bold=True, color=COLOR_STATUS_WARN)
    FONT_STATUS_BAD = Font(bold=True, color=COLOR_STATUS_BAD)
    
    # Fills
    FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    FILL_MATCH = PatternFill(start_color=COLOR_MATCH, end_color=COLOR_MATCH, fill_type="solid")
    FILL_MISMATCH = PatternFill(start_color=COLOR_MISMATCH, end_color=COLOR_MISMATCH, fill_type="solid")
    FILL_MISSING = PatternFill(start_color=COLOR_MISSING, end_color=COLOR_MISSING, fill_type="solid")
    
    # Borders
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alignment
    ALIGN_CENTER = Alignment(horizontal='center')
    ALIGN_LEFT = Alignment(horizontal='left')
    
    # Number formats
    FORMAT_CURRENCY = '€#,##0.00'
    FORMAT_PERCENT = '0.0%'


class ExcelWorkbookBuilder:
    """
    Builds Excel workbooks using Fluent API pattern.
    
    Separates structure creation from styling.
    """
    
    def __init__(self):
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
    
    def create_sheet(self, title: str, position: Optional[int] = None) -> 'SheetBuilder':
        """
        Create a new sheet in the workbook.
        
        Args:
            title: Sheet title
            position: Optional position (0-based index)
            
        Returns:
            SheetBuilder for method chaining
        """
        if position is not None:
            ws = self.wb.create_sheet(title, position)
        else:
            ws = self.wb.create_sheet(title)
        
        return SheetBuilder(ws)
    
    def save(self, path: Path):
        """Save workbook to file."""
        self.wb.save(path)


class SheetBuilder:
    """
    Fluent builder for Excel worksheet content and styling.
    
    Makes Excel generation more maintainable and testable.
    """
    
    def __init__(self, worksheet):
        self.ws = worksheet
    
    def set_title(self, row: int, col: int, title: str, merge_to_col: Optional[int] = None) -> 'SheetBuilder':
        """Add styled title cell."""
        cell = self.ws.cell(row=row, column=col, value=title)
        cell.font = ExcelStyles.FONT_TITLE
        
        if merge_to_col:
            self.ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=merge_to_col
            )
        
        return self
    
    def set_header_row(self, row: int, headers: List[str]) -> 'SheetBuilder':
        """Add styled header row."""
        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_idx, value=header)
            cell.fill = ExcelStyles.FILL_HEADER
            cell.font = ExcelStyles.FONT_HEADER
            cell.border = ExcelStyles.BORDER_THIN
            cell.alignment = ExcelStyles.ALIGN_CENTER
        
        return self
    
    def set_data_row(
        self,
        row: int,
        data: List[Any],
        fill: Optional[PatternFill] = None,
        formats: Optional[Dict[int, str]] = None
    ) -> 'SheetBuilder':
        """
        Add data row with optional styling.
        
        Args:
            row: Row number
            data: List of cell values
            fill: Optional fill color
            formats: Optional dict of {column_index: number_format}
        """
        for col_idx, value in enumerate(data, 1):
            cell = self.ws.cell(row=row, column=col_idx, value=value)
            cell.border = ExcelStyles.BORDER_THIN
            
            if fill:
                cell.fill = fill
            
            if formats and col_idx in formats:
                cell.number_format = formats[col_idx]
        
        return self
    
    def set_cell(
        self,
        row: int,
        col: int,
        value: Any,
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None
    ) -> 'SheetBuilder':
        """Set individual cell with styling."""
        cell = self.ws.cell(row=row, column=col, value=value)
        
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        
        return self
    
    def set_column_widths(self, widths: Dict[str, int]) -> 'SheetBuilder':
        """
        Set column widths.
        
        Args:
            widths: Dict mapping column letter to width
        """
        for col_letter, width in widths.items():
            self.ws.column_dimensions[col_letter].width = width
        
        return self


class ReconciliationExcelPresenter:
    """
    Presents reconciliation data in Excel format.
    
    Pure presentation - no business logic, just formatting.
    """
    
    def create_workbook(self, report_data: Dict[str, Any], output_path: Path):
        """
        Create complete Excel reconciliation report.
        
        Args:
            report_data: Pre-formatted report data (from report generator)
            output_path: Where to save the Excel file
        """
        builder = ExcelWorkbookBuilder()
        
        # Create all sheets
        self._create_summary_sheet(builder, report_data['summary'])
        self._create_matches_sheet(builder, report_data['matches'])
        self._create_mismatches_sheet(builder, report_data['mismatches'])
        self._create_missing_in_pdf_sheet(builder, report_data['missing_in_pdf'])
        self._create_missing_in_evd_sheet(builder, report_data['missing_in_evd'])
        self._create_by_vendor_sheet(builder, report_data['by_vendor'])
        self._create_detailed_comparison_sheet(builder, report_data['detailed'])
        
        # Save
        builder.save(output_path)
    
    def _create_summary_sheet(self, builder: ExcelWorkbookBuilder, data: Dict):
        """Create summary overview sheet."""
        sheet = builder.create_sheet("Summary", 0)
        
        # Title
        sheet.set_title(1, 1, "EVD-PDF Reconciliation Report", merge_to_col=4)
        
        # Metadata
        sheet.set_cell(3, 1, "Report Generated:")
        sheet.set_cell(3, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet.set_cell(4, 1, "Source:")
        sheet.set_cell(4, 2, data.get('source', 'Unknown'))
        
        # Summary statistics
        sheet.set_cell(7, 1, "RECONCILIATION SUMMARY", font=ExcelStyles.FONT_SECTION)
        
        headers = ['Metric', 'Count', 'Percentage']
        sheet.set_header_row(8, headers)
        
        # Data rows
        summary_rows = [
            ['Total EVD Invoices', data['total_evd'], '100%'],
            ['Total PDF Invoices', data['total_pdf'], '100%'],
            ['Perfect Matches', data['matches'], data['match_pct']],
            ['Mismatches', data['mismatches'], data['mismatch_pct']],
            ['Missing in PDF', data['missing_in_pdf'], data['missing_pdf_pct']],
            ['Missing in EVD', data['missing_in_evd'], data['missing_evd_pct']],
        ]
        
        for idx, row_data in enumerate(summary_rows, start=9):
            sheet.set_data_row(idx, row_data)
        
        # Color code specific rows
        sheet.set_cell(11, 2, data['matches'], fill=ExcelStyles.FILL_MATCH)
        sheet.set_cell(12, 2, data['mismatches'], fill=ExcelStyles.FILL_MISMATCH)
        sheet.set_cell(13, 2, data['missing_in_pdf'], fill=ExcelStyles.FILL_MISSING)
        sheet.set_cell(14, 2, data['missing_in_evd'], fill=ExcelStyles.FILL_MISSING)
        
        # Status
        sheet.set_cell(16, 1, "Overall Match Rate:")
        match_rate_font = self._get_status_font(data['match_rate'])
        sheet.set_cell(16, 2, data['match_rate_display'], font=match_rate_font)
        
        sheet.set_cell(17, 1, "Status:")
        status_font = self._get_status_font(data['match_rate'])
        sheet.set_cell(17, 2, data['status_text'], font=status_font)
        
        # Column widths
        sheet.set_column_widths({'A': 30, 'B': 20, 'C': 15, 'D': 15})
    
    def _get_status_font(self, match_rate: float) -> Font:
        """Get appropriate font color based on match rate."""
        if match_rate >= 95:
            return Font(size=14, bold=True, color=ExcelStyles.COLOR_STATUS_GOOD)
        elif match_rate >= 85:
            return Font(size=14, bold=True, color=ExcelStyles.COLOR_STATUS_WARN)
        else:
            return Font(size=14, bold=True, color=ExcelStyles.COLOR_STATUS_BAD)
    
    def _create_matches_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create perfect matches sheet."""
        sheet = builder.create_sheet("Matches")
        
        headers = [
            'Vendor', 'Invoice Number', 'EVD Date', 'PDF Date',
            'EVD Amount (EUR)', 'PDF Amount (EUR)', 'Confidence Score',
            'Amount Note', 'EVD File', 'PDF File'
        ]
        sheet.set_header_row(1, headers)
        
        # Data rows
        for idx, match in enumerate(data, start=2):
            row_data = [
                match['vendor'],
                match['invoice_number'],
                match['evd_date'],
                match['pdf_date'],
                match['evd_amount'],
                match['pdf_amount'],
                match['confidence'],
                match['amount_note'],
                match['evd_file'],
                match['pdf_file']
            ]
            
            sheet.set_data_row(
                idx,
                row_data,
                fill=ExcelStyles.FILL_MATCH,
                formats={5: ExcelStyles.FORMAT_CURRENCY, 6: ExcelStyles.FORMAT_CURRENCY}
            )
        
        sheet.set_column_widths({'A': 25, 'B': 15, 'I': 30, 'J': 30})
    
    def _create_mismatches_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create mismatches sheet."""
        sheet = builder.create_sheet("Mismatches")
        
        headers = [
            'Vendor', 'Invoice Number', 'EVD Amount', 'PDF Amount',
            'Difference', 'Issue', 'EVD File', 'PDF File'
        ]
        sheet.set_header_row(1, headers)
        
        for idx, mismatch in enumerate(data, start=2):
            row_data = [
                mismatch['vendor'],
                mismatch['invoice_number'],
                mismatch['evd_amount'],
                mismatch['pdf_amount'],
                mismatch['difference'],
                mismatch['issue'],
                mismatch['evd_file'],
                mismatch['pdf_file']
            ]
            
            sheet.set_data_row(
                idx,
                row_data,
                fill=ExcelStyles.FILL_MISMATCH,
                formats={3: ExcelStyles.FORMAT_CURRENCY, 4: ExcelStyles.FORMAT_CURRENCY, 5: ExcelStyles.FORMAT_CURRENCY}
            )
        
        sheet.set_column_widths({'A': 25, 'F': 50})
    
    def _create_missing_in_pdf_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create missing in PDF sheet."""
        sheet = builder.create_sheet("Missing in PDF")
        
        headers = ['Vendor', 'Invoice Number', 'Amount', 'Date', 'EVD File']
        sheet.set_header_row(1, headers)
        
        for idx, missing in enumerate(data, start=2):
            row_data = [
                missing['vendor'],
                missing['invoice_number'],
                missing['amount'],
                missing['date'],
                missing['evd_file']
            ]
            
            sheet.set_data_row(
                idx,
                row_data,
                fill=ExcelStyles.FILL_MISSING,
                formats={3: ExcelStyles.FORMAT_CURRENCY}
            )
        
        sheet.set_column_widths({'A': 25, 'E': 35})
    
    def _create_missing_in_evd_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create missing in EVD sheet."""
        sheet = builder.create_sheet("Missing in EVD")
        
        headers = ['Vendor', 'Invoice Number', 'Amount', 'Date', 'PDF File', 'Confidence']
        sheet.set_header_row(1, headers)
        
        for idx, missing in enumerate(data, start=2):
            row_data = [
                missing['vendor'],
                missing['invoice_number'],
                missing['amount'],
                missing['date'],
                missing['pdf_file'],
                missing['confidence']
            ]
            
            sheet.set_data_row(
                idx,
                row_data,
                fill=ExcelStyles.FILL_MISSING,
                formats={3: ExcelStyles.FORMAT_CURRENCY}
            )
        
        sheet.set_column_widths({'A': 25, 'E': 35})
    
    def _create_by_vendor_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create by vendor summary sheet."""
        sheet = builder.create_sheet("By Vendor")
        
        headers = [
            'Vendor', 'EVD Count', 'PDF Count', 'Matches',
            'Mismatches', 'Missing in PDF', 'Missing in EVD', 'Match Rate'
        ]
        sheet.set_header_row(1, headers)
        
        for idx, vendor_data in enumerate(data, start=2):
            row_data = [
                vendor_data['vendor'],
                vendor_data['evd_count'],
                vendor_data['pdf_count'],
                vendor_data['matches'],
                vendor_data['mismatches'],
                vendor_data['missing_pdf'],
                vendor_data['missing_evd'],
                vendor_data['match_rate_display']
            ]
            
            # Color based on match rate
            fill = self._get_match_rate_fill(vendor_data['match_rate'])
            sheet.set_data_row(idx, row_data, fill=fill)
        
        sheet.set_column_widths({'A': 30})
    
    def _get_match_rate_fill(self, match_rate: float) -> PatternFill:
        """Get fill color based on match rate."""
        if match_rate >= 90:
            return ExcelStyles.FILL_MATCH
        elif match_rate >= 70:
            return ExcelStyles.FILL_MISSING
        else:
            return ExcelStyles.FILL_MISMATCH
    
    def _create_detailed_comparison_sheet(self, builder: ExcelWorkbookBuilder, data: List[Dict]):
        """Create detailed comparison sheet."""
        sheet = builder.create_sheet("Detailed Comparison")
        
        headers = [
            'Status', 'Vendor', 'Invoice #', 'EVD Date', 'PDF Date',
            'EVD Amount', 'PDF Amount', 'Difference',
            'EVD Currency', 'PDF Currency', 'Notes'
        ]
        sheet.set_header_row(1, headers)
        
        for idx, row_data_dict in enumerate(data, start=2):
            row_data = [
                row_data_dict['status'],
                row_data_dict['vendor'],
                row_data_dict['invoice_number'],
                row_data_dict['evd_date'],
                row_data_dict['pdf_date'],
                row_data_dict['evd_amount'],
                row_data_dict['pdf_amount'],
                row_data_dict['difference'],
                row_data_dict['evd_currency'],
                row_data_dict['pdf_currency'],
                row_data_dict['notes']
            ]
            
            # Get appropriate fill
            fill = self._get_status_fill(row_data_dict['status'])
            
            sheet.set_data_row(
                idx,
                row_data,
                fill=fill,
                formats={
                    6: ExcelStyles.FORMAT_CURRENCY,
                    7: ExcelStyles.FORMAT_CURRENCY,
                    8: ExcelStyles.FORMAT_CURRENCY
                }
            )
        
        sheet.set_column_widths({'A': 12, 'B': 25, 'K': 50})
    
    def _get_status_fill(self, status: str) -> PatternFill:
        """Get fill color based on status."""
        if status == 'MATCH':
            return ExcelStyles.FILL_MATCH
        elif status == 'MISMATCH':
            return ExcelStyles.FILL_MISMATCH
        else:
            return ExcelStyles.FILL_MISSING
